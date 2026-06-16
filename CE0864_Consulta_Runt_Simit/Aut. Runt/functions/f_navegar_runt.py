
# -*- coding: utf-8 -*-
import sys
import time
import logging
import base64
import re
import os
from datetime import datetime

import openpyxl
import pytesseract
from PIL import Image, ImageEnhance, ImageFilter


from selenium import webdriver
from selenium.webdriver.edge.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException, WebDriverException
from datetime import datetime


# ------------------------------------------------------------------------------ 
# Configuración de Tesseract y logging
# ------------------------------------------------------------------------------ 
pytesseract.pytesseract.tesseract_cmd = r'C:\Tesseract-OCR\tesseract.exe'

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    handlers=[logging.StreamHandler(sys.stdout)]
)
logger = logging.getLogger("main_vim")

# ------------------------------------------------------------------------------ 
# CONSTANTES / SELECTORES (solo VIM/VIN)
# ------------------------------------------------------------------------------ 

# Corregido: usar '>' en vez de '&gt;'
CSS_BOTON_SELECT_CONSULTAR_POR = "mat-form-field.ng-tns-c38-3 > div:nth-child(1) > div:nth-child(1) > div:nth-child(3)"




# Opción VIN por texto visible
XPATH_OPCION_VIN_SPAN = (
    "//mat-option//span[@class='mat-option-text' "
    "and contains(normalize-space(.), 'VIN (Número único de identificación)')]"
)

# Campo VIN por label (robusto)
XPATH_INPUT_NRO_VIN_LABEL = "//mat-form-field[.//mat-label[contains(normalize-space(.), 'Nro. VIN')]]//input"

# Fallback a tu XPath absoluto del mat-form-field
XPATH_MAT_FORMFIELD_VIN = (
    "/html/body/host-runt-root/app-layout/app-theme-runt2/mat-sidenav-container/mat-sidenav-content/div/"
    "ng-component/div/div[2]/div[1]/form/div[2]/div/mat-card/mat-card-content/div[2]/div/div[1]/mat-form-field"
)

# Rutas (ajústalas si cambias tu estructura)
RUTA_CAPTCHA_PNG   = r"C:\RPA\CE0864_Consulta_Runt_Simit\Aut. Runt x VIM\img\captcha.png"
RUTA_EXCEL_IN_OUT  = r"C:\CE0864_Consulta_RUNT_SIMIT\Consulta_RUNT_VIM.xlsx"
RUTA_REG_EJEC      = r"C:\RPA\CE0864_Consulta_Runt_Simit\Aut. Runt\Data\Reg. Ejecuciones\CE0864_RegEjec.xlsx"

URL_RUNT_CONSULTA_VEHICULO = "https://www.runt.gov.co/consultaCiudadana/#/consultaVehiculo"

# ------------------------------------------------------------------------------ 
# Utilidades de archivos/rutas
# ------------------------------------------------------------------------------ 
def ensure_dir_for_file(path_file: str):
    folder = os.path.dirname(path_file)
    if folder and not os.path.exists(folder):
        os.makedirs(folder, exist_ok=True)


def abrir_hoja_excel_vim(ruta_excel: str, nombre_hoja: str = "Listado VIM"):
    """Abrir o crear el libro y hoja de salida para la consulta de VIM/VIN."""
    ensure_dir_for_file(ruta_excel)

    if os.path.exists(ruta_excel):
        wb = openpyxl.load_workbook(ruta_excel)
        if nombre_hoja in wb.sheetnames:
            sh = wb[nombre_hoja]
        else:
            sh = wb.create_sheet(nombre_hoja)
    else:
        wb = openpyxl.Workbook()
        sh = wb.active
        sh.title = nombre_hoja

    # Encabezados esperados
    encabezados = {
        "A1": "VIM",
        "B1": "Estado del SOAT",
        "C1": "Vigencia del SOAT hasta",
        "D1": "Vigencia RTM",
        "E1": "Fecha vigencia RTM",
        "F1": "Fecha expedicion RTM",
        "G1": "Correos",
        "H1": "Fecha de registro"
    }

    for celda, valor in encabezados.items():
        if sh[celda].value != valor:
            sh[celda] = valor

    return wb, sh


def limpiar_datos_salida_vim(sh, fila_inicio=2):
    """Borra resultados previos en las columnas de salida del sheet de VIM."""
    columnas = ['B', 'C', 'D', 'E', 'F', 'H']
    for row in range(fila_inicio, sh.max_row + 1):
        for col in columnas:
            sh[f"{col}{row}"] = None

# ------------------------------------------------------------------------------ 
# Preprocesado de imagen (OCR)
# ------------------------------------------------------------------------------ 
def preprocess_image_strong(image_path):
    image = Image.open(image_path).convert('L')
    image = ImageEnhance.Contrast(image).enhance(3)
    image = image.filter(ImageFilter.MedianFilter(size=3))
    image = image.point(lambda x: 0 if x < 160 else 255)
    image = image.convert('1')
    preprocessed_path = image_path.replace(".png", "_pre.png")
    image.save(preprocessed_path)
    return preprocessed_path

# ------------------------------------------------------------------------------ 
# Captcha (base64) + OCR
# ------------------------------------------------------------------------------ 
def capturar_captcha_base64(driver, ruta_png):
    ensure_dir_for_file(ruta_png)
    captcha = WebDriverWait(driver, 20).until(
        EC.visibility_of_element_located((By.CSS_SELECTOR, 'img.img-responsive.img-fluid'))
    )
    src = captcha.get_attribute("src")
    if not src or not src.startswith("data:image"):
        raise ValueError("El captcha no se encuentra en base64")
    base64_data = re.sub(r"^data:image/.+;base64,", "", src)
    image_bytes = base64.b64decode(base64_data)
    with open(ruta_png, "wb") as f:
        f.write(image_bytes)
    return ruta_png

def ocr_captcha(ruta_png):
    pre = preprocess_image_strong(ruta_png)
    texto = pytesseract.image_to_string(
        pre,
        config='--oem 3 --psm 7 -c tessedit_char_whitelist=ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz0123456789'
    ).strip()
    # Limpieza leve (ej: confusiones comunes)
    texto = texto.replace(" ", "").replace("|", "I").replace("’", "").replace("`", "")
    return texto

# ------------------------------------------------------------------------------ 
# Apertura del select “Consultar por” y elección VIN
# ------------------------------------------------------------------------------ 
def abrir_y_elegir_vin(driver, wait):
    """
    1) Abre el mat-select 'Consultar por' por label (robusto).
    2) Si falla, intenta por id='mat-select-8' y luego por CSS fallback.
    3) Selecciona la opción VIN por su texto visible.
    """
    # Intento por label del mat-form-field
    try:
        # Abrir el select por el trigger dentro del mat-form-field con label "Consultar por"
        field = wait.until(EC.element_to_be_clickable((
            By.XPATH, "//mat-form-field[.//mat-label[contains(normalize-space(.), 'Consultar por')]]"
        )))
        trigger = field.find_element(By.CSS_SELECTOR, ".mat-select-trigger")
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", trigger)
        try:
            trigger.click()
        except Exception:
            driver.execute_script("arguments[0].click();", trigger)
    except Exception:
        # Intento por ID
        try:
            mat_select = wait.until(EC.element_to_be_clickable((By.ID, "mat-select-8")))
            trigger = mat_select.find_element(By.CSS_SELECTOR, ".mat-select-trigger")
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", trigger)
            try:
                trigger.click()
            except Exception:
                driver.execute_script("arguments[0].click();", trigger)
        except Exception:
            # Fallback por CSS (corregido)
            btn_select = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, CSS_BOTON_SELECT_CONSULTAR_POR)))
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", btn_select)
            try:
                btn_select.click()
            except Exception:
                driver.execute_script("arguments[0].click();", btn_select)

    # Esperar el panel del select
    wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, ".cdk-overlay-pane .mat-select-panel")))

    # Seleccionar la opción VIN por texto visible
    opcion_vin_span = wait.until(EC.visibility_of_element_located((By.XPATH, XPATH_OPCION_VIN_SPAN)))
    opcion_vin = opcion_vin_span.find_element(By.XPATH, "./ancestor::mat-option")
    try:
        opcion_vin.click()
    except Exception:
        driver.execute_script("arguments[0].click();", opcion_vin)




def escribir_vin(driver, wait, vin_text):
    """Escribe el VIN directamente en el input con formcontrolname='vin' (del mat-form-field-infix)."""
    try:
        vin_input = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, 'input[formcontrolname="vin"]')))
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", vin_input)
        # Asegurar foco
        try:
            vin_input.click()
        except Exception:
            driver.execute_script("arguments[0].click();", vin_input)

        # Limpiar y escribir
        try:
            vin_input.clear()
        except Exception:
            pass
        vin_input.send_keys(vin_text)

        # Disparar eventos para Angular (input + blur)
        driver.execute_script("""
            const el = arguments[0];
            const val = arguments[1];
            el.value = val;
            el.dispatchEvent(new Event('input', { bubbles: true }));
            el.dispatchEvent(new Event('change', { bubbles: true }));
        """, vin_input, vin_text)
        # Pequeño blur para que se reflejen validaciones
        driver.execute_script("arguments[0].blur();", vin_input)
        time.sleep(0.2)

        # Verificar que el valor quedó
        valor = driver.execute_script("return arguments[0].value;", vin_input)
        if (valor or "").strip() != vin_text.strip():
            raise Exception(f"El VIN no se reflejó en el input (valor leído: '{valor}')")

        logger.info(f"VIN escrito correctamente en formcontrolname='vin': {vin_text}")

    except Exception as e:
        logger.error(f"No se pudo escribir el VIN en formcontrolname='vin': {e}")
        raise




def escribir_captcha(driver, wait, captcha_text):
    """Escribe el captcha por formcontrolname; fallback por label 'Captcha'."""
    try:
        campo_captcha = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, 'input[formcontrolname="captcha"]')))
        campo_captcha.clear()
        campo_captcha.send_keys(captcha_text)
    except Exception:
        input_captcha = wait.until(EC.visibility_of_element_located((
            By.XPATH, "//mat-form-field[.//mat-label[contains(., 'Captcha')]]//input"
        )))
        input_captcha.clear()
        input_captcha.send_keys(captcha_text)

# ------------------------------------------------------------------------------ 
# Click consultar, Swal2 y espera URL
# ------------------------------------------------------------------------------ 
def click_consultar(driver, wait):
    boton = wait.until(EC.element_to_be_clickable(
        (By.XPATH, "//button[.//span[normalize-space()='Consultar Información']]")
    ))
    driver.execute_script("arguments[0].click();", boton)
    return boton

def leer_y_cerrar_swal2(driver, wait):
    try:
        popup = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, ".swal2-popup.swal2-show")))
        classes = popup.get_attribute("class") or ""
        icon = 'info'
        if 'swal2-icon-error' in classes:
            icon = 'error'
        elif 'swal2-icon-warning' in classes:
            icon = 'warning'
        elif 'swal2-icon-success' in classes:
            icon = 'success'
        mensaje_el = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, "#swal2-html-container")))
        mensaje = mensaje_el.text.strip()
        btn_aceptar = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, ".swal2-confirm.swal2-styled")))
        try:
            btn_aceptar.click()
        except Exception:
            driver.execute_script("arguments[0].click();", btn_aceptar)
        return {'mensaje': mensaje, 'icon': icon}
    except TimeoutException:
        return None

def esperar_url_info_vehiculo(driver, timeout=12):
    try:
        WebDriverWait(driver, timeout).until(
            EC.any_of(
                EC.url_contains("#/consulta-vehiculo/consulta/info-vehiculo"),
                EC.url_contains("portalpublico.runt.gov.co/#/consulta-vehiculo/consulta/info-vehiculo")
            )
        )
        return True
    except TimeoutException:
        try:
            href = driver.execute_script("return window.location.href;")
            return (
                "#/consulta-vehiculo/consulta/info-vehiculo" in str(href) or
                "portalpublico.runt.gov.co/#/consulta-vehiculo/consulta/info-vehiculo" in str(href)
            )
        except Exception:
            return False

# ------------------------------------------------------------------------------ 
# Reintentos de consulta por VIM/VIN
# ------------------------------------------------------------------------------ 


from selenium.webdriver.common.keys import Keys


from selenium.webdriver.common.keys import Keys

def intentar_consulta_por_vin(driver, vin_actual, logger, max_reintentos=5):
    wait = WebDriverWait(driver, 15)
    ruta_captcha = RUTA_CAPTCHA_PNG
    popup_timeout = 6
    url_timeout = 12

    try:
        # (1) Abrir menú PREVIO y seleccionar VIN (solo una vez)
        trigger = wait.until(EC.element_to_be_clickable((
            By.XPATH,
            "/html/body/host-runt-root/app-layout/app-theme-runt2/mat-sidenav-container/mat-sidenav-content/div/ng-component/div/div[2]/div[1]/form/div[1]/div/mat-card/mat-card-content/div[2]/div/mat-form-field/div/div[1]/div[3]"
        )))
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", trigger)
        try:
            trigger.click()
        except Exception:
            driver.execute_script("arguments[0].click();", trigger)

        wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, ".cdk-overlay-pane .mat-select-panel")))
        logger.info("Panel del menú previo abierto.")

        opcion_vin_span = wait.until(EC.visibility_of_element_located((
            By.XPATH,
            "//mat-option//span[@class='mat-option-text' and normalize-space()='VIN (Número único de identificación)']"
        )))
        opcion_vin = opcion_vin_span.find_element(By.XPATH, "./ancestor::mat-option")
        try:
            opcion_vin.click()
        except Exception:
            driver.execute_script("arguments[0].click();", opcion_vin)
        logger.info("Opción VIN seleccionada en el menú previo.")

        # Cerrar overlay (evitar bloqueos)
        try:
            driver.switch_to.active_element.send_keys(Keys.ESCAPE)
            time.sleep(0.3)
        except Exception:
            pass

        # (2) (Opcional) Intentar “Consultar por ⇒ VIN” (si falla, seguimos)
       

        # (3) Escribir VIN en el input del mat-form-field-infix (formcontrolname='vin')
        escribir_vin(driver, wait, vin_actual)

    except Exception as e:
        logger.error(f"Error inicial preparando la consulta por VIN: {e}")
        return {"estado": "fallo", "detalle": "No se pudo preparar la consulta por VIN"}

    # (4) Reintentos: SOLO captcha + consulta
    for intento in range(1, max_reintentos + 1):
        logger.info(f"Intento {intento} de captcha para VIN {vin_actual}")

        try:
            # Capturar y resolver captcha
            capturar_captcha_base64(driver, ruta_captcha)
            captcha_text = ocr_captcha(ruta_captcha)
            logger.info(f"Captcha OCR: {captcha_text}")

            # En intentos >= 2, probamos también en mayúsculas (si el sitio las exige)
            captcha_text_alt = captcha_text if intento == 1 else captcha_text

            # Escribir captcha y hacer blur/tab para habilitar botón
            try:
                campo_captcha = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, 'input[formcontrolname="captcha"]')))
                campo_captcha.clear()
                campo_captcha.send_keys(captcha_text_alt)
                campo_captcha.send_keys(Keys.TAB)
                driver.execute_script("arguments[0].blur();", campo_captcha)
            except Exception:
                escribir_captcha(driver, wait, captcha_text_alt)

            # Consultar (click robusto)
            try:
                click_consultar(driver, wait)
            except Exception as e:
                logger.error(f"No se pudo pulsar 'Consultar Información': {e}")
                time.sleep(1)
                continue

            # Validar respuesta: popup o cambio de URL
            try:
                WebDriverWait(driver, popup_timeout).until(
                    EC.visibility_of_element_located((By.CSS_SELECTOR, ".swal2-popup.swal2-show"))
                )
                info = leer_y_cerrar_swal2(driver, wait)
                if info:
                    mensaje = info['mensaje'].strip().lower()
                    icon = info['icon']
                    logger.info(f"SWAL2[{icon}]: {mensaje}")

                    if icon == 'error':
                        if "captcha no es valido" in mensaje or "verifique el valor ingresado" in mensaje:
                            time.sleep(1); continue  # Reintentar captcha
                        elif "no hay información registrada" in mensaje:
                            return {"estado": "sin_info", "detalle": info['mensaje']}
                        elif "no ha sido registrado en el sistema runt" in mensaje or "aún no ha sido registrado" in mensaje:
                            return {"estado": "no_registrado", "detalle": info['mensaje']}
                        else:
                            logger.warning(f"Error Swal2 no contemplado: {info['mensaje']}")
                            time.sleep(1); continue
                    else:
                        if esperar_url_info_vehiculo(driver, timeout=url_timeout):
                            return {"estado": "ok", "detalle": "consulta exitosa (popup info/success + URL)"}
                        logger.info("Popup no-error, pero URL no cambió; reintentando captcha…")
                        time.sleep(1); continue

            except TimeoutException:
                # Sin popup: decidir por URL
                if esperar_url_info_vehiculo(driver, timeout=url_timeout):
                    return {"estado": "ok", "detalle": "consulta exitosa (sin popup + URL)"}
                else:
                    logger.info("Sin popup y URL no cambió; reintentando captcha…")
                    time.sleep(1); continue

        except Exception as e:
            logger.error(f"Error en intento {intento} de captcha: {e}")
            time.sleep(1)
            continue

    # Se agotaron intentos
    return {"estado": "fallo_captcha", "detalle": f"Se agotaron {max_reintentos} intentos sin éxito"}



# ------------------------------------------------------------------------------ 
# Utilidad: leer VIMs desde Excel
# ------------------------------------------------------------------------------ 
def leer_vims_desde_excel(ruta_excel, hoja_idx=0, col_vim="A"):
    wb = openpyxl.load_workbook(ruta_excel)
    sh = wb.worksheets[hoja_idx]
    vims = []
    for row in range(2, sh.max_row + 1):
        v = sh[f"{col_vim}{row}"].value
        if v and str(v).strip():
            vims.append(str(v).strip())
    wb.close()
    return vims

# ------------------------------------------------------------------------------ 
# Flujo principal (SOLO VIM/VIN)
# ------------------------------------------------------------------------------ 
def f_navegar_runt_vim():
    try:
        global cont_procesados, cont_excepciones, cont_exitosos
        cont_procesados = 0
        cont_excepciones = 0
        cont_exitosos = 0

        inicio = datetime.now()
        logger.info("Obteniendo la lista de VIM/VIN")

        # === Leer VIMs ===
        try:
            ruta_input = RUTA_EXCEL_IN_OUT
            vim_list = leer_vims_desde_excel(ruta_input, hoja_idx=0, col_vim="A")
            # Limpiar resultados anteriores manteniendo la lista de VIMs en la columna A.
            wb_temp, sh_temp = abrir_hoja_excel_vim(ruta_input)
            limpiar_datos_salida_vim(sh_temp)
            wb_temp.save(ruta_input)
        except Exception as e:
            logger.error(f"No fue posible obtener VIMs: {e}")
            return

        fila_actual = 2  # fila inicial en Excel

        for vim_actual in vim_list:
            cont_procesados += 1
            logger.info(f"Procesando VIM/VIN={vim_actual}")
            print("Procesando VIM/VIN:", vim_actual)

            # Configurar navegador (opcional: headless)
            options = Options()
            # options.add_argument('--headless')  # descomenta para ejecución sin GUI
            driver = webdriver.Edge(options=options)
            wait = WebDriverWait(driver, 20)

            try:
                # Navegar directo (evitar requests.get)
                driver.get(URL_RUNT_CONSULTA_VEHICULO)
                # Pequeña espera para Angular
                time.sleep(5)

                # Selección VIN + consulta (con reintentos)
                resultado = intentar_consulta_por_vin(driver, vim_actual, logger, max_reintentos=5)

                # === Escritura en Excel de salida ===
                ruta_excel_salida = RUTA_EXCEL_IN_OUT
                try:
                    wb_out, sh_out = abrir_hoja_excel_vim(ruta_excel_salida)
                except Exception as e:
                    print("Error al abrir/crear el Excel de salida:", e)
                    logger.error("Error al abrir/crear el Excel de salida", exc_info=True)
                    cont_excepciones += 1
                    continue

                if resultado["estado"] == "ok":
                    # Abrir sección SOAT
                    try:
                        headers = WebDriverWait(driver, 10).until(
                            EC.presence_of_all_elements_located(
                                (By.CSS_SELECTOR, "mat-expansion-panel-header.mat-expansion-panel-header.mat-focus-indicator")
                            )
                        )
                        objetivo = None
                        for h in headers:
                            try:
                                titulo_el = h.find_element(By.CSS_SELECTOR, ".mat-expansion-panel-header-title")
                                titulo = (titulo_el.text or "").strip().lower()
                            except Exception:
                                titulo = (h.text or "").strip().lower()
                            if "soat" in titulo or "póliza soat" in titulo:
                                objetivo = h; break

                        if not objetivo:
                            raise Exception("No se encontró un panel cuyo título contenga 'SOAT'.")

                        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", objetivo)
                        WebDriverWait(driver, 10).until(EC.element_to_be_clickable(objetivo))
                        if objetivo.get_attribute("aria-expanded") != "true":
                            try:
                                objetivo.click()
                            except Exception:
                                driver.execute_script("arguments[0].click();", objetivo)
                            WebDriverWait(driver, 5).until(lambda d: objetivo.get_attribute("aria-expanded") == "true")

                        controls_id = objetivo.get_attribute("aria-controls")
                        if controls_id:
                            WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.ID, controls_id)))
                            try:
                                WebDriverWait(driver, 10).until(
                                    EC.visibility_of_element_located((By.XPATH, f"//*[@id='{controls_id}']//mat-row[1]"))
                                )
                            except Exception:
                                pass
                    except Exception as e:
                        print("Error al abrir la sección SOAT:", e)
                        logger.error("Error al abrir la sección SOAT: %s", e)
                        cont_excepciones += 1

                   
                    # Estado SOAT (col B)
                    estado_texto = "Sin info."
                    try:
                        estado_cell = WebDriverWait(driver, 10).until(
                            EC.visibility_of_element_located((By.CSS_SELECTOR, "mat-row:nth-of-type(1) mat-cell.mat-column-estado"))
                        )

                        # Extraer el texto del estado sin el <mat-icon>
                        texto_sin_icono = driver.execute_script("""
                            const el = arguments[0].cloneNode(true);
                            el.querySelectorAll('mat-icon, svg, i').forEach(n => n.remove());
                            return el.textContent.trim();
                        """, estado_cell)

                        # Normaliza espacios y mayúsculas (opcional)
                        estado_texto = " ".join(texto_sin_icono.split()).upper()  # -> 'VIGENTE', 'VENCIDO', etc.

                    except Exception as e:
                        print("Error al copiar el estado del SOAT:", e)
                        logger.error("Error al copiar el estado del SOAT", exc_info=True)
                        cont_excepciones += 1
                    
                    # Fecha fin de vigencia SOAT (col C)
                    fecha_v = "Sin info."
                    try:
                        vigencia_cell = WebDriverWait(driver, 12).until(
                            EC.visibility_of_element_located((By.CSS_SELECTOR, "mat-row:nth-of-type(1) mat-cell.mat-column-fechaFinVigencia"))
                        )
                        fecha_v = driver.execute_script("return arguments[0].textContent;", vigencia_cell).strip()
                    except Exception as e:
                        print("Error al copiar la fecha de vigencia del SOAT:", e)
                        logger.error("Error al copiar la fecha de vigencia del SOAT", exc_info=True)
                        cont_excepciones += 1
                        
                        
                    
                    # --- FECHA DE REGISTRO (col I) — del div con ícono calendario
                    try:
                        fecha_registro_el = WebDriverWait(driver, 10).until(
                            EC.visibility_of_element_located((By.CSS_SELECTOR, "div.td-icon-date b"))
                        )
                        fecha_registro = (fecha_registro_el.text or "").strip()
                        print("Fecha Registro:", fecha_registro)
                    except TimeoutException:
                        print("No se encontró el div de Fecha Registro (td-icon-date)")
                        logger.error("No se encontró el div de Fecha Registro")
                        cont_excepciones += 1


                    # Abrir RTM
                    try:
                        rtm = WebDriverWait(driver, 10).until(
                            EC.element_to_be_clickable((
                                By.XPATH,
                                "//mat-expansion-panel-header[.//mat-panel-title[normalize-space()='Certificado de revisión técnico mecánica y de emisiones contaminantes (RTM)']]"
                            ))
                        )
                        rtm.click(); time.sleep(1)
                    except Exception as e:
                        print("Error al encontrar el botón RTM en el RUNT")
                        logger.error("Error al encontrar el botón RTM en el RUNT")
                        cont_excepciones += 1

                    # Vigencia RTM (col D)
                    vige_rtm = "Sin info."
                    try:
                        vigencia_rtm = WebDriverWait(driver, 10).until(
                            EC.visibility_of_element_located((By.CSS_SELECTOR, "mat-row:nth-of-type(1) mat-cell.mat-column-vigente"))
                        )
                        vige_rtm = vigencia_rtm.text.strip()
                    except Exception as e:
                        print("Error al copiar la vigencia del RTM en el RUNT")
                        logger.error("Error al copiar la vigencia del RTM en el RUNT")
                        cont_excepciones += 1

                    # Fecha vigencia RTM (col E)
                    f_r_r = "Sin info."
                    try:
                        fecha_vigencia_rtm = WebDriverWait(driver, 10).until(
                            EC.visibility_of_element_located((By.CSS_SELECTOR, "mat-row:nth-of-type(1) mat-cell.mat-column-fechaVigencia"))
                        )
                        f_r_r = fecha_vigencia_rtm.text.strip()
                    except Exception as e:
                        print("Error al encontrar y copiar la fecha de vigencia del RTM en el RUNT")
                        logger.error("Error al encontrar y copiar la fecha de vigencia del RTM en el RUNT")
                        cont_excepciones += 1

                    # Fecha expedición RTM (col F)
                    f_e_r = "Sin info."
                    try:
                        fecha_exp_cell = WebDriverWait(driver, 10).until(EC.visibility_of_element_located((
                            By.XPATH,
                            ("//mat-row[.//mat-cell[contains(@class,'mat-column-tipoRevision') and "
                             "normalize-space()='REVISION TECNICO-MECANICO']]"
                             "//mat-cell[contains(@class,'mat-column-fechaExpedicion')]")
                        )))
                        f_e_r = fecha_exp_cell.text.strip()
                    except Exception as e:
                        print("Error al copiar la fecha de expedición del RTM en el RUNT:", e)
                        logger.error("Error al copiar la fecha de expedición del RTM en el RUNT")
                        cont_excepciones += 1
                        
                    
                    # Guardar en Excel salida (A: VIM ya existente; B..F datos; G: Correos se deja como esté)
                    try:
                        sh_out[f"A{fila_actual}"] = vim_actual
                        sh_out[f"B{fila_actual}"] = (estado_texto or "Sin info")
                        sh_out[f"C{fila_actual}"] = (fecha_v or "Sin info")
                        sh_out[f"D{fila_actual}"] = (vige_rtm or "Sin info")
                        sh_out[f"E{fila_actual}"] = (f_r_r or "Sin info")
                        sh_out[f"F{fila_actual}"] = (f_e_r or "Sin info")
                        sh_out[f"H{fila_actual}"] = (fecha_registro or "Sin info")   # ✅ Columna H confirmada
                        wb_out.save(ruta_excel_salida)
                    except PermissionError:
                        temp_path = ruta_excel_salida.replace(".xlsx", "_temp.xlsx")
                        wb_out.save(temp_path)
                        print(f"El archivo estaba bloqueado. Se guardó copia temporal en: {temp_path}")
                        logger.warning(f"Excel bloqueado. Copia temporal: {temp_path}")
                    except Exception as e:
                        print("Error al pegar datos en el Excel:", e)
                        logger.error("Error al pegar datos en el Excel", exc_info=True)
                        cont_excepciones += 1

                    cont_exitosos += 1

                elif resultado["estado"] == "sin_info":
                    for col in ['B', 'C', 'D', 'E', 'F', "H"]:
                        sh_out[col + str(fila_actual)] = "Sin info"
                    try:
                        wb_out.save(ruta_excel_salida)
                    except Exception as e:
                        print("Error al escribir 'Sin info' en el Excel:", e)
                        logger.error("Error al escribir 'Sin info' en el Excel")

                elif resultado["estado"] in ("mismatch", "fallo_captcha", "no_registrado"):
                    valor = "Error" if resultado["estado"] != "no_registrado" else "No registrado"
                    for col in ['B', 'C', 'D', 'E', 'F' , "H"]:
                        sh_out[col + str(fila_actual)] = valor
                    try:
                        wb_out.save(ruta_excel_salida)
                    except Exception as e:
                        print("Error al escribir estado en el Excel:", e)
                        logger.error("Error al escribir estado en el Excel")

            finally:
                time.sleep(1)
                fila_actual += 1
                try:
                    driver.quit()
                except Exception:
                    pass

        # Fin del for
        fin = datetime.now()

        # Registro de ejecución
        try:
            ensure_dir_for_file(RUTA_REG_EJEC)
            workbook1 = openpyxl.load_workbook(RUTA_REG_EJEC)
            hoja = workbook1.active
            fila = hoja.max_row + 1
            ejecucion_id = inicio.strftime("%Y%m%d%H%M%S")
            hoja['A' + str(fila)] = ejecucion_id
            hoja['B' + str(fila)] = cont_procesados
            hoja['C' + str(fila)] = cont_excepciones
            hoja['D' + str(fila)] = cont_exitosos
            hoja['E' + str(fila)] = 0
            hoja['F' + str(fila)] = inicio.strftime("%Y-%m-%d %H:%M:%S")
            hoja['G' + str(fila)] = fin.strftime("%Y-%m-%d %H:%M:%S")
            workbook1.save(RUTA_REG_EJEC)
        except Exception as e:
            print("Error al guardar el archivo de registro", e)

    except Exception as e:
        print("Error al ejecutar la función f_navegar_runt_vim:", e)
        logger.error("Error al ejecutar la función f_navegar_runt_vim", exc_info=True)
        try:
            cont_excepciones += 1
        except Exception:
            pass

# ------------------------------------------------------------------------------ 
# Wrapper
# ------------------------------------------------------------------------------ 
def f_navegar_runt():
    return f_navegar_runt_vim()

# ------------------------------------------------------------------------------ 
# Ejecución directa
# ------------------------------------------------------------------------------ 
if __name__ == "__main__":
    f_navegar_runt_vim()

